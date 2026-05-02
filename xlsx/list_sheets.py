import openpyxl

wb = openpyxl.load_workbook('D:\\MyCodeSpace\\Linux_Knowledge\\Embedded_Knowledge_System_V2026\\xlsx\\Linux知识体系.xlsx', read_only=True, data_only=True)
with open('sheets.txt', 'w', encoding='utf-8') as f:
    for i, name in enumerate(wb.sheetnames):
        f.write(f'{i}: {name}\n')
        ws = wb[name]
        f.write(f'   rows={ws.max_row}, cols={ws.max_column}\n')
