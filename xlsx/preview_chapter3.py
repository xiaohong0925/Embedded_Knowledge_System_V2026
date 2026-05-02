import openpyxl
from openpyxl.utils import get_column_letter
import json
import os

xlsx_path = r'D:\MyCodeSpace\Linux_Knowledge\Embedded_Knowledge_System_V2026\xlsx\Linux知识体系.xlsx'
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

print("All sheets:")
for i, name in enumerate(wb.sheetnames):
    print(f"  {i}: {name}")

# Find sheet for chapter 3
chapter3_sheet = None
for i, name in enumerate(wb.sheetnames):
    if 'Linux内核' in name or '内核深度' in name or '3' in name:
        chapter3_sheet = wb[name]
        print(f"\nFound Chapter 3 sheet: '{name}' at index {i}")
        break

if not chapter3_sheet:
    print("No chapter 3 sheet found by name, checking all sheets...")
    for i, name in enumerate(wb.sheetnames):
        ws = wb[name]
        if ws.max_row > 0:
            first_val = ws.cell(1, 1).value
            if first_val and ('Linux' in str(first_val) or '内核' in str(first_val) or '3.' in str(first_val)):
                chapter3_sheet = ws
                print(f"Found Chapter 3 sheet by content: '{name}'")
                break

if chapter3_sheet:
    ws = chapter3_sheet
    print(f"\nSheet dimensions: rows={ws.max_row}, cols={ws.max_column}")
    
    # Extract structure
    data = []
    for row in range(1, min(10, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val:
                row_data.append(f"Col{col}: {str(val)[:100]}")
        if row_data:
            print(f"Row {row}: {' | '.join(row_data)}")

print("\nDone preview.")
