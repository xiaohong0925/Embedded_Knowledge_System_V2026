import openpyxl
from openpyxl.styles import PatternFill, Font
import json
import os

xlsx_path = r'D:\MyCodeSpace\Linux_Knowledge\Embedded_Knowledge_System_V2026\xlsx\Linux知识体系.xlsx'
output_dir = r'D:\MyCodeSpace\Linux_Knowledge\Embedded_Knowledge_System_V2026\xlsx\chapter3_extract'
os.makedirs(output_dir, exist_ok=True)

wb = openpyxl.load_workbook(xlsx_path, data_only=False)

# Sheet index 10 is "linux内核" (第3章 Linux内核深度解析)
ws = wb.worksheets[10]
print(f"Processing sheet: '{wb.sheetnames[10]}'")
print(f"Dimensions: rows={ws.max_row}, cols={ws.max_column}")

# Extract all data
data = {
    'sheet_name': wb.sheetnames[10],
    'title': None,
    'sections': []
}

current_section = None
current_subsection = None
current_topic = None

for row_idx in range(1, ws.max_row + 1):
    # Read key columns
    col1 = ws.cell(row_idx, 1).value  # H1/H2 title
    col2 = ws.cell(row_idx, 2).value  # H3 title
    col3 = ws.cell(row_idx, 3).value  # Description/summary
    col8 = ws.cell(row_idx, 8).value  # Content start
    col9 = ws.cell(row_idx, 9).value  # Main content
    col10 = ws.cell(row_idx, 10).value  # Extended content
    col11 = ws.cell(row_idx, 11).value  # Additional content
    col12 = ws.cell(row_idx, 12).value
    col13 = ws.cell(row_idx, 13).value
    col14 = ws.cell(row_idx, 14).value  # Extended/back-to-index
    
    # Check for font color
    def get_color(cell):
        if cell.font and cell.font.color and cell.font.color.rgb:
            return str(cell.font.color.rgb)
        return None
    
    color_info = {
        'col1': get_color(ws.cell(row_idx, 1)),
        'col2': get_color(ws.cell(row_idx, 2)),
        'col9': get_color(ws.cell(row_idx, 9)),
        'col10': get_color(ws.cell(row_idx, 10)),
    }
    
    # Title row
    if row_idx == 1 and col1:
        data['title'] = str(col1).strip()
        continue
    
    # Skip empty rows
    has_content = any([col1, col2, col3, col8, col9, col10, col11, col12, col13, col14])
    if not has_content:
        continue
    
    # Determine level by column content
    entry = {
        'row': row_idx,
        'col1': str(col1).strip() if col1 else None,
        'col2': str(col2).strip() if col2 else None,
        'col3': str(col3).strip() if col3 else None,
        'col8': str(col8).strip() if col8 else None,
        'col9': str(col9).strip() if col9 else None,
        'col10': str(col10).strip() if col10 else None,
        'col11': str(col11).strip() if col11 else None,
        'col12': str(col12).strip() if col12 else None,
        'col13': str(col13).strip() if col13 else None,
        'col14': str(col14).strip() if col14 else None,
        'colors': {k:v for k,v in color_info.items() if v},
    }
    
    # Categorize by level
    if col1 and not col2 and not col3:
        # Likely H1 section header
        current_section = {
            'level': 'H1',
            'title': str(col1).strip(),
            'topics': [],
            'raw': entry
        }
        data['sections'].append(current_section)
        current_subsection = None
    elif col2 and not col1:
        # Likely H2/H3 topic header
        if current_section is None:
            # Orphan topic, create a section
            current_section = {
                'level': 'H1_orphan',
                'title': 'Unnamed Section',
                'topics': [],
            }
            data['sections'].append(current_section)
        
        current_subsection = {
            'level': 'H2/H3',
            'title': str(col2).strip(),
            'content': [],
            'raw': entry
        }
        current_section['topics'].append(current_subsection)
    else:
        # Content row
        content_entry = {
            'content_columns': {},
            'raw': entry
        }
        for col_name, val in [('col8', col8), ('col9', col9), ('col10', col10), ('col11', col11), ('col12', col12), ('col13', col13), ('col14', col14)]:
            if val:
                content_entry['content_columns'][col_name] = str(val).strip()
        
        if current_subsection:
            current_subsection['content'].append(content_entry)
        elif current_section:
            current_section.setdefault('floating_content', []).append(content_entry)

# Save JSON
json_path = os.path.join(output_dir, 'chapter3_structure.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"\nJSON saved to: {json_path}")
print(f"Total sections: {len(data['sections'])}")
for s in data['sections']:
    print(f"  Section: {s.get('title', 'N/A')} ({len(s.get('topics', []))} topics)")

# Also generate a markdown summary
md_path = os.path.join(output_dir, 'chapter3_structure.md')
with open(md_path, 'w', encoding='utf-8') as f:
    f.write(f"# {data.get('title', 'Chapter 3')}\n\n")
    for s in data['sections']:
        f.write(f"## {s.get('title', '')}\n\n")
        for t in s.get('topics', []):
            f.write(f"### {t.get('title', '')}\n\n")
            for c in t.get('content', []):
                for col, val in c.get('content_columns', {}).items():
                    f.write(f"**{col}**:\n{val}\n\n")
            f.write("---\n\n")

print(f"Markdown saved to: {md_path}")
print("Done.")
