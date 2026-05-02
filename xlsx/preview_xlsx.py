import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\15314\.openclaw\workspace\downloads\19de3995-86a2-8846-8000-00004ea5fac3_Linux知识体系.xlsx', read_only=True, data_only=True)
ws = wb['驱动模型']

with open(r'C:\Users\15314\.kimi_openclaw\workspace\xlsx_preview.txt', 'w', encoding='utf-8') as f:
    f.write(f'Total rows: {ws.max_row}, cols: {ws.max_column}\n\n')
    
    # Print header row
    header = [str(v)[:30] if v else '' for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    f.write(f'Header: {header}\n\n')
    
    # Print first 50 rows with all columns
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        vals = []
        for i, v in enumerate(row):
            if v:
                s = str(v)[:80].replace('\n', '\\n')
                vals.append(f'Col{i}:{s}')
        if vals:
            f.write(f'Row {row_idx}: ' + ' | '.join(vals) + '\n')

print('done')
