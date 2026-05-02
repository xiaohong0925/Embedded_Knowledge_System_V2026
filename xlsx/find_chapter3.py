import openpyxl
import json
import os

xlsx_path = r'D:\MyCodeSpace\Linux_Knowledge\Embedded_Knowledge_System_V2026\xlsx\Linux知识体系.xlsx'
wb = openpyxl.load_workbook(xlsx_path, data_only=False)

# Check all sheets for chapter 3 content
chapter3_sheets = []

for idx, name in enumerate(wb.sheetnames):
    if 'WpsReserved' in name:
        continue
    ws = wb[name]
    if ws.max_row < 3:
        continue
    
    # Check first few rows for chapter indicators
    for row in range(1, min(5, ws.max_row + 1)):
        for col in range(1, min(16, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val:
                s = str(val)
                if '3.' in s or 'Linux内核深度' in s or '内核深度' in s:
                    chapter3_sheets.append({
                        'index': idx,
                        'name': name,
                        'found_in': f'row={row},col={col}',
                        'text': s[:100]
                    })
                    break
        else:
            continue
        break

print(f"Found {len(chapter3_sheets)} potential Chapter 3 sheets:")
for s in chapter3_sheets:
    print(f"  [{s['index']}] '{s['name']}': {s['text'][:80]}...")

# Also check for sheets that might be sub-chapters of 3
# Look for specific section names
subchapters = ['内存管理', '虚拟文件系统', '设备树', '驱动模型', '中断', '时间子系统', 
               '电源管理', '网络子系统', '内核安全', '进程管理', '调度']

print("\nAll sheets with possible Chapter 3 sub-chapters:")
for idx, name in enumerate(wb.sheetnames):
    if 'WpsReserved' in name:
        continue
    for sub in subchapters:
        if sub in name:
            ws = wb[name]
            print(f"  [{idx}] '{name}' (rows={ws.max_row}, cols={ws.max_column})")
            break

# Print ALL sheets with their first row content
print("\n=== ALL SHEETS (first row preview) ===")
for idx, name in enumerate(wb.sheetnames):
    if 'WpsReserved' in name:
        continue
    ws = wb[name]
    first_vals = []
    for col in range(1, min(16, ws.max_column + 1)):
        v = ws.cell(1, col).value
        if v:
            first_vals.append(f'C{col}:{str(v)[:40]}')
    print(f"[{idx:2d}] '{name}' (r={ws.max_row},c={ws.max_column}): {' | '.join(first_vals[:3])}")
