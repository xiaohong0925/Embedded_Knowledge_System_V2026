import openpyxl
from openpyxl.styles import PatternFill, Font
import json
import os

xlsx_path = r'D:\MyCodeSpace\Linux_Knowledge\Embedded_Knowledge_System_V2026\xlsx\Linux知识体系.xlsx'
output_dir = r'D:\MyCodeSpace\Linux_Knowledge\Embedded_Knowledge_System_V2026\xlsx\chapter3_extract'
os.makedirs(output_dir, exist_ok=True)

wb = openpyxl.load_workbook(xlsx_path, data_only=False)

# Chapter 3 sheets mapping
chapter3_sheets = {
    12: '3.1 进程管理与调度',
    13: '3.2 内存管理',
    14: '3.3 虚拟文件系统与页缓存',
    16: '3.4 设备树 & ACPI',
    17: '3.5 驱动模型',
    18: '3.6 中断子系统',
    19: '3.7 时间子系统',
}

def extract_sheet_v2(ws, sheet_idx, sheet_label):
    """Improved extraction with better hierarchy detection"""
    data = {
        'sheet_index': sheet_idx,
        'sheet_name': wb.sheetnames[sheet_idx],
        'sheet_label': sheet_label,
        'rows': ws.max_row,
        'cols': ws.max_column,
        'structure': []
    }
    
    current_h1 = None
    current_h2 = None
    
    for row_idx in range(1, ws.max_row + 1):
        # Read key columns
        def get_cell_value(col):
            cell = ws.cell(row_idx, col)
            if cell.value:
                return str(cell.value).strip()
            return None
        
        col1 = get_cell_value(1)   # H1/H2 title
        col2 = get_cell_value(2)   # H3 or description
        col3 = get_cell_value(3)   # Content/description
        col4 = get_cell_value(4)
        col8 = get_cell_value(8)
        col9 = get_cell_value(9)
        col10 = get_cell_value(10)
        col11 = get_cell_value(11)
        col12 = get_cell_value(12)
        col13 = get_cell_value(13)
        col14 = get_cell_value(14)
        
        # Skip empty rows
        if not any([col1, col2, col3, col4, col8, col9, col10, col11, col12, col13, col14]):
            continue
        
        # Collect all content
        content = {}
        for col_idx, val in [(3, col3), (4, col4), (8, col8), (9, col9), (10, col10), (11, col11), (12, col12), (13, col13), (14, col14)]:
            if val:
                content[f'col{col_idx}'] = val
        
        # Determine level
        # H1: Only col1, no content in other columns, or just col14 (back to index)
        is_h1 = col1 and (not col3 and not col9 and not col10)
        
        # H2: col1 + content in col3/col9/col10, or col2 present
        is_h2 = (col1 and (col3 or col9 or col10 or col11)) or (col2 and not col1)
        
        # Content: no col1/col2, only content columns
        is_content = not col1 and not col2 and content
        
        if is_h1:
            current_h1 = {
                'level': 'H1',
                'title': col1,
                'row': row_idx,
                'children': []
            }
            data['structure'].append(current_h1)
            current_h2 = None
        elif is_h2:
            if current_h1 is None:
                current_h1 = {
                    'level': 'H1_orphan',
                    'title': '未命名章节',
                    'row': row_idx,
                    'children': []
                }
                data['structure'].append(current_h1)
            
            title = col2 if col2 else (col1 if col1 else '未命名')
            current_h2 = {
                'level': 'H2',
                'title': title,
                'row': row_idx,
                'content': content
            }
            current_h1['children'].append(current_h2)
        elif is_content:
            if current_h2:
                # Merge content into current H2
                for k, v in content.items():
                    current_h2['content'][k] = v
            elif current_h1:
                current_h1.setdefault('floating_content', []).append({
                    'row': row_idx,
                    'content': content
                })
    
    return data

all_data = {
    'chapter': '3. Linux内核深度解析',
    'source_file': xlsx_path,
    'sheets': []
}

for idx, label in chapter3_sheets.items():
    if idx >= len(wb.sheetnames):
        continue
    sheet_name = wb.sheetnames[idx]
    ws = wb[sheet_name]
    print(f"Extracting sheet [{idx}] '{sheet_name}' -> {label}...")
    
    sheet_data = extract_sheet_v2(ws, idx, label)
    all_data['sheets'].append(sheet_data)
    
    h1_count = len(sheet_data['structure'])
    h2_count = sum(len(h1.get('children', [])) for h1 in sheet_data['structure'])
    content_rows = sum(
        sum(len(h2.get('content', {})) for h2 in h1.get('children', []))
        for h1 in sheet_data['structure']
    )
    print(f"  H1: {h1_count}, H2: {h2_count}, content items: {content_rows}")

# Save JSON
json_path = os.path.join(output_dir, 'chapter3_complete_v2.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"\nJSON saved: {json_path}")

# Generate detailed Markdown
md_path = os.path.join(output_dir, 'chapter3_complete_v2.md')
with open(md_path, 'w', encoding='utf-8') as f:
    f.write(f"# {all_data['chapter']}\n\n")
    f.write("> 数据来源：xlsx原始数据提取\n")
    f.write(f"> 文件：{os.path.basename(xlsx_path)}\n")
    f.write(f"> 提取时间：2026-05-01\n\n")
    
    for sheet in all_data['sheets']:
        f.write(f"---\n\n")
        f.write(f"## {sheet['sheet_label']} (Sheet: {sheet['sheet_name']})\n\n")
        
        for h1 in sheet['structure']:
            f.write(f"### {h1['title']}\n\n")
            
            for h2 in h1.get('children', []):
                f.write(f"#### {h2['title']}\n\n")
                
                # Write content columns
                for col_name, val in h2.get('content', {}).items():
                    f.write(f"**{col_name}**:\n\n")
                    f.write(f"{val}\n\n")
                
                f.write("\n")
            
            # Floating content
            for fc in h1.get('floating_content', []):
                f.write(f"\n> **浮动内容** (row {fc['row']}):\n\n")
                for col_name, val in fc['content'].items():
                    f.write(f"**{col_name}**: {val[:300]}...\n\n")

print(f"Markdown saved: {md_path}")

# Also create a compact JSON for quick reference
compact = {
    'chapter': all_data['chapter'],
    'sheets_summary': []
}
for sheet in all_data['sheets']:
    sections = []
    for h1 in sheet['structure']:
        topics = []
        for h2 in h1.get('children', []):
            # Count content length
            total_len = sum(len(v) for v in h2.get('content', {}).values())
            topics.append({
                'title': h2['title'],
                'content_chars': total_len
            })
        sections.append({
            'section': h1['title'],
            'topics': topics
        })
    compact['sheets_summary'].append({
        'label': sheet['sheet_label'],
        'sections': sections
    })

compact_path = os.path.join(output_dir, 'chapter3_summary.json')
with open(compact_path, 'w', encoding='utf-8') as f:
    json.dump(compact, f, ensure_ascii=False, indent=2)

print(f"Summary JSON saved: {compact_path}")
print("Done.")
