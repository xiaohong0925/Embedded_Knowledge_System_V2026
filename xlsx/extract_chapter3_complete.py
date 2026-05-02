import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import json
import os
import re

xlsx_path = r'D:\MyCodeSpace\Linux_Knowledge\Embedded_Knowledge_System_V2026\xlsx\Linux知识体系.xlsx'
output_dir = r'D:\MyCodeSpace\Linux_Knowledge\Embedded_Knowledge_System_V2026\xlsx\chapter3_extract'
os.makedirs(output_dir, exist_ok=True)

wb = openpyxl.load_workbook(xlsx_path, data_only=False)

# Chapter 3 sheets mapping (based on analysis)
chapter3_sheets = {
    12: '3.1 进程管理与调度',
    13: '3.2 内存管理',
    14: '3.3 虚拟文件系统与页缓存',
    16: '3.4 设备树 & ACPI',
    17: '3.5 驱动模型',
    18: '3.6 中断子系统',
    19: '3.7 时间子系统',
}

# Additional sheets that might belong to ch3
# 4: 时钟&电源管理 -> could be 1.4 or 3.7
# 33: 网络编程 -> 7.2 or 3.8
# 36: 网络编程与系统层 -> could be 3.8
# 38: 系统安全 -> could be 3.9 or 8.x

def extract_sheet(ws, sheet_idx, sheet_name):
    """Extract complete structure from a sheet"""
    data = {
        'sheet_index': sheet_idx,
        'sheet_name': sheet_name,
        'rows': ws.max_row,
        'cols': ws.max_column,
        'structure': []
    }
    
    current_h1 = None
    current_h2 = None
    current_h3 = None
    
    for row_idx in range(1, ws.max_row + 1):
        # Read all columns
        row_data = {}
        has_content = False
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            val = cell.value
            if val:
                has_content = True
                row_data[f'col{col_idx}'] = {
                    'value': str(val),
                    'font_color': str(cell.font.color.rgb) if cell.font and cell.font.color and cell.font.color.rgb else None,
                    'bold': cell.font.bold if cell.font else None,
                    'fill_color': str(cell.fill.fgColor.rgb) if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb else None,
                }
        
        if not has_content:
            continue
        
        # Determine hierarchy level
        col1 = row_data.get('col1', {}).get('value', '')
        col2 = row_data.get('col2', {}).get('value', '')
        col3 = row_data.get('col3', {}).get('value', '')
        
        entry = {
            'row': row_idx,
            'columns': row_data
        }
        
        # Heuristic: if col1 has content but col2 doesn't, it's H1
        # If col2 has content, it's H2/H3
        # If only col3+, it's content
        
        is_h1 = col1 and not col2 and (not col3 or len(col3) < 200)
        is_h2 = col2 and not col1
        
        if is_h1:
            current_h1 = {
                'level': 'H1',
                'title': col1.strip(),
                'row': row_idx,
                'children': []
            }
            data['structure'].append(current_h1)
            current_h2 = None
            current_h3 = None
        elif is_h2:
            if current_h1 is None:
                # Orphan H2
                current_h1 = {
                    'level': 'H1_orphan',
                    'title': '未命名章节',
                    'children': []
                }
                data['structure'].append(current_h1)
            
            current_h2 = {
                'level': 'H2',
                'title': col2.strip(),
                'row': row_idx,
                'content': []
            }
            current_h1['children'].append(current_h2)
            current_h3 = None
        else:
            # Content row - aggregate all content columns
            content_cols = {}
            for k, v in row_data.items():
                if k.startswith('col') and int(k[3:]) >= 3:
                    content_cols[k] = v
            
            if content_cols:
                content_entry = {
                    'level': 'content',
                    'row': row_idx,
                    'content': content_cols
                }
                
                if current_h2:
                    current_h2['content'].append(content_entry)
                elif current_h1:
                    current_h1.setdefault('floating_content', []).append(content_entry)
                else:
                    # Orphan content
                    data.setdefault('orphan_content', []).append(content_entry)
    
    return data

# Extract all chapter 3 sheets
all_data = {
    'chapter': '3. Linux内核深度解析',
    'sheets': []
}

for idx, label in chapter3_sheets.items():
    if idx >= len(wb.sheetnames):
        continue
    sheet_name = wb.sheetnames[idx]
    ws = wb[sheet_name]
    print(f"Extracting sheet [{idx}] '{sheet_name}' -> {label}...")
    
    sheet_data = extract_sheet(ws, idx, label)
    all_data['sheets'].append(sheet_data)
    
    # Print summary
    h1_count = len(sheet_data['structure'])
    h2_count = sum(len(h1.get('children', [])) for h1 in sheet_data['structure'])
    print(f"  H1: {h1_count}, H2: {h2_count}, rows: {sheet_data['rows']}")

# Save JSON
json_path = os.path.join(output_dir, 'chapter3_complete.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"\nJSON saved: {json_path}")

# Generate Markdown summary
md_path = os.path.join(output_dir, 'chapter3_complete.md')
with open(md_path, 'w', encoding='utf-8') as f:
    f.write(f"# {all_data['chapter']}\n\n")
    f.write("> 自动生成目录，基于xlsx原始数据提取\n\n")
    
    for sheet in all_data['sheets']:
        f.write(f"## {sheet['sheet_name']}\n\n")
        
        for h1 in sheet['structure']:
            f.write(f"### {h1['title']}\n\n")
            
            for h2 in h1.get('children', []):
                f.write(f"#### {h2['title']}\n\n")
                
                # Write content
                for content_entry in h2.get('content', []):
                    for col_name, col_data in content_entry['content'].items():
                        val = col_data['value']
                        # Truncate long content
                        if len(val) > 500:
                            val = val[:500] + "... [truncated]"
                        f.write(f"\n**{col_name}**: {val}\n\n")
                
                f.write("---\n\n")
            
            # Write floating content
            for fc in h1.get('floating_content', []):
                for col_name, col_data in fc['content'].items():
                    val = col_data['value']
                    if len(val) > 500:
                        val = val[:500] + "... [truncated]"
                    f.write(f"\n**{col_name}** (floating): {val}\n\n")

print(f"Markdown saved: {md_path}")
print("Done.")
