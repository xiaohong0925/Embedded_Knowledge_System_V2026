import openpyxl
import json
import os

xlsx_path = r'D:\MyCodeSpace\Linux_Knowledge\Embedded_Knowledge_System_V2026\xlsx\Linux知识体系.xlsx'
output_dir = r'D:\MyCodeSpace\Linux_Knowledge\Embedded_Knowledge_System_V2026\xlsx\prep_extract'
os.makedirs(output_dir, exist_ok=True)

wb = openpyxl.load_workbook(xlsx_path, data_only=False)

# Sheet mapping for my tasks
target_sheets = {
    # Chapter 6: 内核调试与性能优化
    29: '6.1 内核崩溃分析',
    30: '6.2 动态追踪技术', 
    31: '6.3 实时性优化',
    # Chapter 1: 硬件层
    1: '1.1 CPU架构',
    2: '1.2 总线协议',
    4: '1.3 时钟与电源管理',
    5: '1.4 存储设备',
    6: '1.5 硬件调试',
    7: '1.6 硬件安全',
    # Network subsystem
    33: '3.10 网络子系统',
    36: '7.2 网络编程',
}

def extract_sheet(ws, sheet_idx, label):
    """Extract structure from a sheet"""
    data = {
        'index': sheet_idx,
        'label': label,
        'name': wb.sheetnames[sheet_idx],
        'rows': ws.max_row,
        'cols': ws.max_column,
        'structure': []
    }
    
    current_section = None
    
    for row_idx in range(1, ws.max_row + 1):
        col1 = ws.cell(row_idx, 1).value
        col2 = ws.cell(row_idx, 2).value
        col3 = ws.cell(row_idx, 3).value
        col9 = ws.cell(row_idx, 9).value
        col10 = ws.cell(row_idx, 10).value
        col11 = ws.cell(row_idx, 11).value
        col14 = ws.cell(row_idx, 14).value
        
        has_content = any([col1, col2, col3, col9, col10, col11, col14])
        if not has_content:
            continue
        
        def to_str(v):
            return str(v).strip() if v else None
        
        entry = {
            'row': row_idx,
            'col1': to_str(col1),
            'col2': to_str(col2),
            'col3': to_str(col3),
            'col9': to_str(col9),
            'col10': to_str(col10),
            'col11': to_str(col11),
            'col14': to_str(col14),
        }
        
        # Remove None values
        entry = {k: v for k, v in entry.items() if v is not None}
        
        if entry.get('col1') and not entry.get('col3') and not entry.get('col9'):
            # H1 header
            current_section = {
                'level': 'H1',
                'title': entry['col1'],
                'row': row_idx,
                'children': []
            }
            data['structure'].append(current_section)
        elif entry.get('col1') and (entry.get('col3') or entry.get('col9') or entry.get('col10')):
            # H2 with content
            if current_section is None:
                current_section = {
                    'level': 'H1_orphan',
                    'title': 'Unnamed',
                    'children': []
                }
                data['structure'].append(current_section)
            
            child = {
                'level': 'H2',
                'title': entry.get('col1', ''),
                'content': {}
            }
            for k in ['col3', 'col9', 'col10', 'col11', 'col14']:
                if k in entry:
                    child['content'][k] = entry[k]
            current_section['children'].append(child)
        elif entry.get('col2') and not entry.get('col1'):
            # H3
            if current_section is None:
                current_section = {
                    'level': 'H1_orphan',
                    'title': 'Unnamed',
                    'children': []
                }
                data['structure'].append(current_section)
            
            child = {
                'level': 'H2/H3',
                'title': entry['col2'],
                'content': {}
            }
            for k in ['col3', 'col9', 'col10', 'col11', 'col14']:
                if k in entry:
                    child['content'][k] = entry[k]
            current_section['children'].append(child)
    
    return data

all_data = []
for idx, label in target_sheets.items():
    if idx >= len(wb.sheetnames):
        print(f"SKIP: index {idx} out of range")
        continue
    ws = wb.worksheets[idx]
    print(f"Extracting [{idx}] '{wb.sheetnames[idx]}' -> {label}...")
    sheet_data = extract_sheet(ws, idx, label)
    all_data.append(sheet_data)
    
    h1 = len(sheet_data['structure'])
    h2 = sum(len(s.get('children', [])) for s in sheet_data['structure'])
    content = sum(
        sum(len(v) for v in c.get('content', {}).values())
        for s in sheet_data['structure']
        for c in s.get('children', [])
    )
    print(f"  H1={h1}, H2={h2}, content_chars={content}")

# Save JSON
json_path = os.path.join(output_dir, 'task_sheets.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"\nSaved: {json_path}")

# Generate markdown summary
md_path = os.path.join(output_dir, 'task_sheets.md')
with open(md_path, 'w', encoding='utf-8') as f:
    f.write("# 任务相关Sheet提取结果\n\n")
    f.write("提取范围：第1章硬件层 + 第6章内核调试 + 网络子系统\n\n")
    
    for sheet in all_data:
        f.write(f"---\n\n## {sheet['label']} (Sheet: {sheet['name']})\n\n")
        f.write(f"- 行数: {sheet['rows']}, 列数: {sheet['cols']}\n")
        f.write(f"- H1节: {len(sheet['structure'])}\n\n")
        
        for h1 in sheet['structure']:
            f.write(f"### {h1['title']}\n\n")
            for h2 in h1.get('children', []):
                f.write(f"#### {h2['title']}\n\n")
                for col, val in h2.get('content', {}).items():
                    f.write(f"**{col}**: {val[:500]}\n\n")
                f.write("\n")

print(f"Saved: {md_path}")
print("Done.")
